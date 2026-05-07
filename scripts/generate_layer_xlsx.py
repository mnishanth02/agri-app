from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Common styles
title_font = Font(name='Arial', size=16, bold=True)
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
data_font = Font(name='Arial', size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ─── Sheet 1: Summary ───
ws = wb.active
ws.title = "Summary"

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 42
ws.column_dimensions['D'].width = 42
ws.column_dimensions['E'].width = 32
ws.column_dimensions['F'].width = 38

ws['A1'] = "Crop Viz — Map Layer Architecture"
ws['A1'].font = title_font
ws.merge_cells('A1:F1')

ws['A2'] = "4 layers rendered top-to-bottom on a single MapLibre GL JS map canvas"
ws['A2'].font = Font(name='Arial', size=11, italic=True, color='666666')
ws.merge_cells('A2:F2')

headers = ['#', 'Layer', 'Library', 'Service / Provider', 'Free Tier', 'Beyond Free Tier']
header_fill = PatternFill('solid', fgColor='343A40')

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

rows = [
    ['4 (top)', 'Field polygon', 'terra-draw + MapLibre GL JS', 'None — your own data', '$0 always', '$0 always'],
    ['3', 'NDVI / dpRVI heatmap', 'MapLibre GL JS (image source)', 'Copernicus Data Space — Sentinel Hub', '10,000 PU/month', 'Commercial plan (contact Copernicus)'],
    ['2', 'Roads + place labels', 'MapLibre GL JS + @esri/maplibre-arcgis', 'ArcGIS Location Platform (Basemap Styles)', 'Shared with Layer 1', 'Shared with Layer 1'],
    ['1 (bottom)', 'Satellite photo', 'MapLibre GL JS + @esri/maplibre-arcgis', 'ArcGIS Location Platform (Basemap Styles)', '1,000 sessions/mo OR 2M tiles/mo', '$4/1,000 sessions OR $0.15/1,000 tiles'],
]

row_colors = ['D0BFFF', 'FFEC99', 'B2F2BB', 'A5D8FF']

for i, (row_data, color) in enumerate(zip(rows, row_colors), 5):
    fill = PatternFill('solid', fgColor=color)
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.font = data_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws['A10'] = "Key Insights"
ws['A10'].font = Font(name='Arial', size=12, bold=True)
ws.merge_cells('A10:F10')

insights = [
    "• Layers 1 & 2 share one ArcGIS session/tile budget — switching from imagery/standard to imagery adds labels at no extra cost",
    "• Session model (Oct 2025) is recommended — 12hr unlimited tiles per session, billed per user visit",
    "• Layer 3 (NDVI) is the core product differentiator — 100 acres × 4 scans/yr uses only ~0.13 PU of 10,000 free",
    "• Layer 4 is entirely free open-source with no external dependencies",
]
for i, insight in enumerate(insights, 11):
    ws.cell(row=i, column=1, value=insight).font = Font(name='Arial', size=10)
    ws.merge_cells(f'A{i}:F{i}')

# ─── Sheet 2: All Layers (2×2 grid layout) ───
ws2 = wb.create_sheet("All Layers - Detail")

# Column widths: Left panel (A-C) | Gap (D) | Right panel (E-G)
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 44
ws2.column_dimensions['C'].width = 5
ws2.column_dimensions['D'].width = 20
ws2.column_dimensions['E'].width = 44


def write_layer_block(sheet, start_row, start_col, layer_num, position_label, title, description, color, bg_color, tech_items, pricing_items):
    """Write a compact layer block at the specified position."""
    r = start_row
    col_key = start_col
    col_val = start_col + 1

    # Layer header row with colored background
    header_fill = PatternFill('solid', fgColor=color)
    cell = sheet.cell(row=r, column=col_key, value=f"LAYER {layer_num} ({position_label})")
    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.border = thin_border
    cell2 = sheet.cell(row=r, column=col_val, value=title)
    cell2.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell2.fill = header_fill
    cell2.border = thin_border
    r += 1

    # Description
    light_fill = PatternFill('solid', fgColor=bg_color)
    cell = sheet.cell(row=r, column=col_key, value="Purpose")
    cell.font = Font(name='Arial', size=9, bold=True)
    cell.fill = light_fill
    cell.border = thin_border
    cell2 = sheet.cell(row=r, column=col_val, value=description)
    cell2.font = Font(name='Arial', size=9, italic=True)
    cell2.fill = light_fill
    cell2.border = thin_border
    cell2.alignment = Alignment(wrap_text=True)
    r += 1

    # Technical details
    for key, val in tech_items:
        sheet.cell(row=r, column=col_key, value=key).font = Font(name='Arial', size=9, bold=True)
        sheet.cell(row=r, column=col_key).border = thin_border
        sheet.cell(row=r, column=col_val, value=val).font = Font(name='Arial', size=9)
        sheet.cell(row=r, column=col_val).border = thin_border
        sheet.cell(row=r, column=col_val).alignment = Alignment(wrap_text=True)
        r += 1

    # Pricing section
    price_fill = PatternFill('solid', fgColor=bg_color)
    for key, val in pricing_items:
        sheet.cell(row=r, column=col_key, value=key).font = Font(name='Arial', size=9, bold=True)
        sheet.cell(row=r, column=col_key).fill = price_fill
        sheet.cell(row=r, column=col_key).border = thin_border
        sheet.cell(row=r, column=col_val, value=val).font = Font(name='Arial', size=9)
        sheet.cell(row=r, column=col_val).fill = price_fill
        sheet.cell(row=r, column=col_val).border = thin_border
        sheet.cell(row=r, column=col_val).alignment = Alignment(wrap_text=True)
        r += 1

    return r


# Title
ws2['A1'] = "All Layers — Detailed Technical Reference"
ws2['A1'].font = Font(name='Arial', size=14, bold=True)
ws2.merge_cells('A1:E1')
ws2['A2'] = "2×2 layout: Layer 4 & 3 (top row) · Layer 2 & 1 (bottom row)"
ws2['A2'].font = Font(name='Arial', size=10, italic=True, color='868E96')
ws2.merge_cells('A2:E2')

# ── TOP-LEFT: Layer 4 ──
write_layer_block(ws2, 4, 1, 4, "TOP", "Field Polygon (User-Drawn)",
    "Boundary agronomist draws around field — white outline on top",
    "6741D9", "F3EDFF",
    [
        ("Drawing lib", "terra-draw + terra-draw-maplibre-gl-adapter (MIT)"),
        ("Rendering", "MapLibre GL JS — geojson source, fill + line layers"),
        ("Data format", "GeoJSON Feature<Polygon>"),
        ("Storage", "React state (proto) → PostgreSQL + PostGIS (prod)"),
        ("External svc", "None — your own app's data and code"),
    ],
    [
        ("💰 terra-draw", "Free — MIT open source"),
        ("💰 MapLibre", "Free — BSD 3-Clause"),
        ("💰 Total", "$0 always"),
    ])

# ── TOP-RIGHT: Layer 3 ──
write_layer_block(ws2, 4, 4, 3, "CORE VALUE", "NDVI / dpRVI Satellite Overlay",
    "Crop health heatmap — Sentinel-2 optical + Sentinel-1 SAR (works through clouds)",
    "E67700", "FFF9DB",
    [
        ("Service", "Copernicus Data Space — Sentinel Hub API"),
        ("NDVI source", "Sentinel-2 L2A — 10m res, B04+B08 bands"),
        ("dpRVI source", "Sentinel-1 SAR GRD — C-band, 5–20m res"),
        ("India revisit", "S-2: 5 days (3-sat) · S-1: 6–12 days"),
        ("API type", "WMS → PNG tile for polygon bbox + date"),
        ("WMS URL", "sh.dataspace.copernicus.eu/ogc/wms/{ID}"),
        ("Rendering", "MapLibre GL JS — image source"),
        ("Account", "Free at dataspace.copernicus.eu"),
    ],
    [
        ("💰 Free tier", "10,000 PU/month"),
        ("💰 1 PU =", "512×512 px, 3 bands, 16-bit"),
        ("💰 100 ac × 1 scan", "≈ 0.03 PU"),
        ("💰 100 ac × 4/yr", "≈ 0.13 PU (well within free!)"),
    ])

# ── BOTTOM-LEFT: Layer 2 ──
row_bottom = 20
write_layer_block(ws2, row_bottom, 1, 2, "MIDDLE", "Roads + Place Labels Overlay",
    "Village/district names, roads, water labels — location context",
    "2B8A3E", "EBFBEE",
    [
        ("How to use", "Switch style to arcgis/imagery (labels bundled)"),
        ("Data source", "Esri vector tiles — roads, places, boundaries"),
        ("Alternative", "ArcGIS Open Basemaps (Aug 2025) — Overture+OSM"),
        ("Rendering", "MapLibre GL JS + @esri/maplibre-arcgis"),
        ("Attribution", "Auto-handled by plugin (Esri ToS)"),
    ],
    [
        ("💰 Separate cost?", "No — included in Layer 1 quota"),
        ("💰 Tile tier", "Shared 2M tiles/mo with Layer 1"),
        ("💰 Session", "Same session covers both layers"),
    ])

# ── BOTTOM-RIGHT: Layer 1 ──
write_layer_block(ws2, row_bottom, 4, 1, "BOTTOM", "Satellite Imagery Background",
    "Photo of Earth — terrain, crops, rivers. Base for agronomist navigation.",
    "1864AB", "E7F5FF",
    [
        ("Style", "arcgis/imagery/standard (photo, no labels)"),
        ("Imagery", "Maxar Vivid — 30cm urban, 50cm agri, 2.5m global"),
        ("India res", "30–50cm — field bunds visible at zoom 16"),
        ("Rendering", "MapLibre GL JS v4.x (BSD 3-Clause, free)"),
        ("Plugin", "@esri/maplibre-arcgis (Apache 2.0, free)"),
        ("Service", "ArcGIS Location Platform — Basemap Styles"),
        ("Account", "Free signup at developers.arcgis.com"),
    ],
    [
        ("💰 Option A", "Tiles: 2M free/mo → $0.15/1K tiles"),
        ("💰 Option B ←", "Sessions: 1K free/mo → $4/1K sessions"),
        ("💰 Session =", "12hr unlimited tiles (recommended)"),
        ("💰 Note", "Oct 2025 — per visit, not per pan/zoom"),
    ])

# ─── Sheet 3: Services Overview ───
ws3 = wb.create_sheet("Services Overview")
ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 32
ws3.column_dimensions['D'].width = 38

ws3['A1'] = "External Services Overview"
ws3['A1'].font = Font(name='Arial', size=14, bold=True)
ws3.merge_cells('A1:D1')

ws3['A2'] = "All external service dependencies, pricing tiers, and account requirements"
ws3['A2'].font = Font(name='Arial', size=10, italic=True, color='868E96')
ws3.merge_cells('A2:D2')

svc_headers = ['Service', 'Layers', 'Free Tier', 'Paid Tier']
svc_fill = PatternFill('solid', fgColor='495057')

for col, h in enumerate(svc_headers, 1):
    cell = ws3.cell(row=4, column=col, value=h)
    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell.fill = svc_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

svc_data = [
    ('ArcGIS Location Platform', '1 & 2', '1K sessions OR 2M tiles/mo', '$4/1K sessions · $0.15/1K tiles'),
    ('Copernicus Sentinel Hub', '3', '10,000 PU/month', 'Commercial plan (contact Copernicus)'),
    ('None (open source only)', '4', '$0 always', '$0 always'),
]

svc_colors = ['D0EBFF', 'FFF3BF', 'E8D5F5']

for i, (row_data, color) in enumerate(zip(svc_data, svc_colors), 5):
    fill = PatternFill('solid', fgColor=color)
    for col, val in enumerate(row_data, 1):
        cell = ws3.cell(row=i, column=col, value=val)
        cell.font = Font(name='Arial', size=10)
        cell.fill = fill
        cell.border = thin_border

# ArcGIS detail section
ws3['A9'] = "ArcGIS Location Platform — Detail"
ws3['A9'].font = Font(name='Arial', size=12, bold=True, color='1864AB')
ws3.merge_cells('A9:D9')

arcgis_details = [
    ('Service URL', 'location.arcgis.com', 'Signup', 'developers.arcgis.com (free)'),
    ('Covers', 'Satellite imagery + Roads/Labels', 'Style (photo)', 'arcgis/imagery/standard'),
    ('Style (labels)', 'arcgis/imagery', 'Plugin', '@esri/maplibre-arcgis (Apache 2.0)'),
    ('Tile model', '2M tiles/mo free', 'Overage', '$0.15 per 1,000 tiles'),
    ('Session model ←', '1,000 sessions/mo free', 'Overage', '$4.00 per 1,000 sessions'),
    ('Session duration', '12 hours, unlimited tiles', 'Billing', 'Per user visit, not per pan/zoom'),
]

for i, row_data in enumerate(arcgis_details, 10):
    for col, val in enumerate(row_data, 1):
        cell = ws3.cell(row=i, column=col, value=val)
        cell.font = Font(name='Arial', size=9)
        cell.border = thin_border
        if col in (1, 3):
            cell.font = Font(name='Arial', size=9, bold=True)

# Copernicus detail section
ws3['A17'] = "Copernicus Sentinel Hub — Detail"
ws3['A17'].font = Font(name='Arial', size=12, bold=True, color='E67700')
ws3.merge_cells('A17:D17')

cop_details = [
    ('Service URL', 'dataspace.copernicus.eu', 'Dashboard', 'Sentinel Hub Dashboard'),
    ('Covers', 'NDVI (Sentinel-2) + dpRVI (Sentinel-1)', 'API', 'WMS endpoint'),
    ('Free tier', '10,000 PU/month', '1 PU =', '512×512 px, 3 bands, 16-bit'),
    ('100 ac × 1 scan', '≈ 0.03 PU', '100 ac × 4/yr', '≈ 0.13 PU'),
    ('Vienna 413 km² yr', '886 PU', 'Conclusion', 'Agri use well within free tier'),
]

for i, row_data in enumerate(cop_details, 18):
    for col, val in enumerate(row_data, 1):
        cell = ws3.cell(row=i, column=col, value=val)
        cell.font = Font(name='Arial', size=9)
        cell.border = thin_border
        if col in (1, 3):
            cell.font = Font(name='Arial', size=9, bold=True)

output_path = r'c:\Users\v-mnmurugan\projects\GIS\projects\demo-project-crop-viz\agri-app\docs\map-layer-architecture.xlsx'
wb.save(output_path)
print(f"Saved: {output_path}")
